import os
import argparse
import pandas as pd
from openpyxl import load_workbook, Workbook
import glob
import psycopg2

print('Establishing Database Connection...')

def main(args):
    # Create connection string from the command-line arguments
    conn_string = f"postgresql://{args.user}:{args.password}@{args.host}:{args.port}/{args.db_name}"

    # Iterate over xlsx files in the specified directory
    for filename in os.listdir(args.dir_path):
        if filename.endswith(".xlsx"):
            process_xlsx_file(filename, args.dir_path, conn_string)

def process_xlsx_file(filename, dir_path, conn_string):
    # Extract prefix from the filename for dynamic SQL query construction
    prefix = filename[:3]
    sqlquery = construct_sql_query(prefix)

    # Execute the query and fetch the results using psycopg2
    try:
        with psycopg2.connect(conn_string) as conn:
            df = pd.read_sql_query(sqlquery, conn)
    except (Exception, psycopg2.Error) as e:
        print(f"Error: {e}. Ensure the query is correct.")
        df = pd.DataFrame()

    if not df.empty:
        append_data_to_workbook(filename, dir_path, prefix, df)

def construct_sql_query(prefix):
    # Construct a generic SQL query
    sqlquery = f"""
        SELECT ...
        FROM your_database.your_table
        WHERE your_column LIKE UPPER('{prefix}%')
    """
    return sqlquery

def append_data_to_workbook(filename, dir_path, prefix, df):
    # Load workbook or create a new one if it doesn't exist
    wb_path = os.path.join(dir_path, filename)
    if os.path.exists(wb_path):
        wb = load_workbook(wb_path)
    else:
        wb = Workbook()
        wb.create_sheet("Your_Sheet_Name")

    # Append Data to the specified sheet
    sheet_name = "Your_Sheet_Name"
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    sheet = wb[sheet_name]
    sheet.append(df.columns.tolist())
    for row in df.itertuples(index=False):
        sheet.append(list(row))

    output_filename = os.path.join(dir_path, f"{prefix}_Generic_Name.xlsx")
    wb.save(output_filename)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Process xlsx files and update them based on SQL query results.")
    parser.add_argument("dir_path", help="Path to the directory containing the xlsx files.")
    parser.add_argument("--user", required=True, help="Database user.")
    parser.add_argument("--password", required=True, help="Database password.")
    parser.add_argument("--host", required=True, help="Database host.")
    parser.add_argument("--port", required=True, help="Database port.")
    parser.add_argument("--db_name", required=True, help="Database name.")

    args = parser.parse_args()
    main(args)